"""Build a progress-tracking XLSX workbook from a workout plan JSON file."""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Font, PatternFill

from plan_model import default_output_paths, iter_exercise_rows, load_plan


PLAN_HEADERS = [
    "Week",
    "Day",
    "Focus",
    "Block",
    "Exercise",
    "Visual/source",
    "Sets x reps/time",
    "Intensity",
    "Rest",
    "Notes",
]

LOG_HEADERS = [
    "Date",
    "Week",
    "Day",
    "Block",
    "Exercise",
    "Planned",
    "Load",
    "Completed reps/time",
    "RPE",
    "RIR",
    "Pain flag",
    "Status",
    "Notes",
]


def style_header(row) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in row:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize_columns(sheet, max_width: int = 42) -> None:
    for column in sheet.columns:
        letter = column[0].column_letter
        width = 10
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(max_width, len(value) + 2))
        sheet.column_dimensions[letter].width = width


def add_table(sheet, name: str, end_row: int, end_col: str, start_row: int = 1) -> None:
    table = Table(displayName=name, ref=f"A{start_row}:{end_col}{max(end_row, start_row + 1)}")
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    sheet.add_table(table)


def build_xlsx(plan: dict[str, Any], output_path: str | Path, log_rows: int = 300) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    plan_sheet = workbook.active
    plan_sheet.title = "Plan"
    log_sheet = workbook.create_sheet("Log")
    summary_sheet = workbook.create_sheet("Summary")

    plan_sheet.append(PLAN_HEADERS)
    style_header(plan_sheet[1])
    plan_rows = []
    for row in iter_exercise_rows(plan):
        plan_rows.append(
            [
                row.get("week", ""),
                row.get("day", ""),
                row.get("focus", ""),
                row.get("block", ""),
                row.get("exercise", ""),
                row.get("visual_source", ""),
                row.get("sets_reps", ""),
                row.get("intensity", ""),
                row.get("rest", ""),
                row.get("notes", ""),
            ]
        )
    for item in plan_rows:
        plan_sheet.append(item)

    for cell in plan_sheet["F"][1:]:
        if cell.value:
            cell.hyperlink = str(cell.value)
            cell.value = "source"
            cell.style = "Hyperlink"

    plan_sheet.freeze_panes = "A2"
    add_table(plan_sheet, "PlanTable", len(plan_rows) + 1, "J")

    log_sheet.append(LOG_HEADERS)
    style_header(log_sheet[1])
    for index in range(log_rows):
        plan_index = (index % max(len(plan_rows), 1)) + 2
        row_num = index + 2
        log_sheet.append(
            [
                "",
                f"=IF(Plan!A{plan_index}=\"\",\"\",Plan!A{plan_index})",
                f"=IF(Plan!B{plan_index}=\"\",\"\",Plan!B{plan_index})",
                f"=IF(Plan!D{plan_index}=\"\",\"\",Plan!D{plan_index})",
                f"=IF(Plan!E{plan_index}=\"\",\"\",Plan!E{plan_index})",
                f"=IF(Plan!G{plan_index}=\"\",\"\",Plan!G{plan_index})",
                "",
                "",
                "",
                "",
                "No",
                "Planned",
                "",
            ]
        )

    log_sheet.freeze_panes = "A2"
    add_table(log_sheet, "LogTable", log_rows + 1, "M")

    status_validation = DataValidation(type="list", formula1='"Planned,Done,Skipped,Modified"', allow_blank=False)
    pain_validation = DataValidation(type="list", formula1='"No,Yes"', allow_blank=False)
    log_sheet.add_data_validation(status_validation)
    log_sheet.add_data_validation(pain_validation)
    status_validation.add(f"L2:L{log_rows + 1}")
    pain_validation.add(f"K2:K{log_rows + 1}")

    log_sheet.conditional_formatting.add(
        f"K2:K{log_rows + 1}",
        CellIsRule(operator="equal", formula=['"Yes"'], fill=PatternFill("solid", fgColor="F4CCCC")),
    )
    log_sheet.conditional_formatting.add(
        f"L2:L{log_rows + 1}",
        CellIsRule(operator="equal", formula=['"Done"'], fill=PatternFill("solid", fgColor="D9EAD3")),
    )

    summary_sheet["A1"] = plan["title"]
    summary_sheet["A1"].font = Font(size=16, bold=True, color="1F4E78")
    summary_sheet["A2"] = f"Generated from plan date: {plan['date']}"
    metrics = [
        ("Planned log rows", f"=COUNTA(Log!E2:E{log_rows + 1})"),
        ("Completed rows", f'=COUNTIF(Log!L2:L{log_rows + 1},"Done")'),
        ("Skipped rows", f'=COUNTIF(Log!L2:L{log_rows + 1},"Skipped")'),
        ("Modified rows", f'=COUNTIF(Log!L2:L{log_rows + 1},"Modified")'),
        ("Completion rate", '=IF(B5=0,"",B6/B5)'),
        ("Average RPE", f'=IFERROR(AVERAGEIF(Log!I2:I{log_rows + 1},">0"),"")'),
        ("Pain flags", f'=COUNTIF(Log!K2:K{log_rows + 1},"Yes")'),
    ]
    summary_sheet.append([])
    summary_sheet.append(["Metric", "Value"])
    style_header(summary_sheet[4])
    for metric in metrics:
        summary_sheet.append(list(metric))
    summary_sheet["B9"].number_format = "0%"
    summary_sheet.freeze_panes = "A4"
    add_table(summary_sheet, "SummaryTable", len(metrics) + 4, "B", start_row=4)

    for sheet in [plan_sheet, log_sheet, summary_sheet]:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        autosize_columns(sheet)

    workbook.save(output)
    return output


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build a progress tracker XLSX from JSON.")
    parser.add_argument("plan_json", help="Path to workout plan JSON.")
    parser.add_argument("--output", help="Output XLSX path. Defaults to programms/<date>-<slug>-progress.xlsx.")
    parser.add_argument("--output-dir", default="programms", help="Directory used when --output is omitted.")
    parser.add_argument("--log-rows", type=int, default=300, help="Number of editable log rows to prebuild.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    plan = load_plan(args.plan_json)
    _, default_xlsx = default_output_paths(plan, args.output_dir)
    output = Path(args.output) if args.output else default_xlsx
    path = build_xlsx(plan, output, log_rows=args.log_rows)
    print(path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
